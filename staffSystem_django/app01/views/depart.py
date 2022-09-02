# -*- coding:utf-8 -*-
# @Time : 2022/8/25 20:42
# @Author: fbz
# @File : depart.py
from django.shortcuts import render, redirect, HttpResponse
from openpyxl import load_workbook

from app01.models import *
from app01.utils.pagination import Pagination


def depart_list(request):
    """ 部门列表 """
    queryset = Department.objects.all()

    page_object = Pagination(request, queryset)

    context = {
        'queryset': page_object.page_queryset,  # 分页数据
        'page_string': page_object.html(),  # 页码
    }

    return render(request, "depart_list.html", context)


def depart_add(request):
    """ 添加部门 """
    if request.method == "GET":
        return render(request, "depart_add.html")

    # 获取用户post提交过来的数据
    title = request.POST.get("title")

    # 保存到数据库
    Department.objects.create(title=title)

    # 重定向回部门列表
    return redirect("/depart/list/")


def depart_delete(request):
    """ 删除部门 """
    # 获取ID
    nid = request.GET.get('nid')
    # 删除
    Department.objects.filter(id=nid).delete()

    # 重定向回部门列表
    return redirect("/depart/list/")


def depart_edit(request, nid):
    """ 修改部门 """
    if request.method == 'GET':
        # 根据nid获取数据[obj,]
        row_object = Department.objects.filter(id=nid).first()

        return render(request, "depart_edit.html", {'row_object': row_object})

    # 获取用户提交的标题
    title = request.POST.get('title')
    # 根据id找到数据库中数据更新
    Department.objects.filter(id=nid).update(title=title)

    # 重定向回部门列表
    return redirect("/depart/list/")


def depart_multi(request):
    """ 批量上传(Excel文件) """

    # 获取用户上传的文件对象
    file_object = request.FILES.get('exc')
    # print(type(file_object))

    # from openpyxl import load_workbook
    # 对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    # cell = sheet.cell(1, 1)  # 第一行第一列的cell对象
    # print(cell.value)  # cell的值

    # 循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        # print(row)
        text = row[0].value
        # print(text)
        exists = Department.objects.filter(title=text).exists()
        if not exists:
            Department.objects.create(title=text)

    return redirect('/depart/list/')
